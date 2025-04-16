import openpyxl
import tkinter as tk
from tkinter import filedialog, messagebox
from pathlib import Path
import os
from dotenv import load_dotenv

# Carga de variables de entorno (.env) y extracción de categorías permitidas
load_dotenv()
# Ejemplo en el archivo .env: ProductCategory=Desktops,Laptops,Monitors,Workstations
allowed_categories = [s.strip() for s in os.getenv("ProductCategory", "").split(",") if s.strip()]

def log_message(message):
    """Agrega un mensaje al widget de logs y lo imprime en consola."""
    print(message)
    log_text.insert(tk.END, message + "\n")
    log_text.see(tk.END)

def open_file():
    global source_file
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if not file_path:
        log_message("No se seleccionó ningún archivo fuente.")
        return
    source_file = Path(file_path)
    log_message(f"Archivo fuente seleccionado: {source_file}")

def select_destination():
    global destination_file
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if not file_path:
        log_message("No se seleccionó ningún archivo destino.")
        return
    destination_file = Path(file_path)
    log_message(f"Archivo destino seleccionado: {destination_file}")
    messagebox.showinfo("Destino Seleccionado", f"Archivo destino:\n{destination_file}")

def copy_data():
    """
    Copia los datos del archivo fuente al destino respetando:
    - Sólo se copian filas cuyo valor en la columna A se encuentre en allowed_categories.
    - Columna S (índice 18) se copia a la columna E del destino.
    - Columna L (índice 11) se copia a la columna H del destino.
    - Se inicia en la primera fila vacía disponible en el destino.
    - Se evita copiar duplicados de Part Number (columna C, índice 2).
    - Se omiten filas cuyo precio (columna L) sea 0.
    """
    if not source_file or not destination_file:
        messagebox.showerror("Error", "Por favor, selecciona ambos archivos: fuente y destino.")
        log_message("Error: Falta seleccionar archivo fuente y/o destino.")
        return

    # Abrir archivo fuente
    try:
        log_message("Abriendo archivo fuente...")
        wb_source = openpyxl.load_workbook(source_file, data_only=True)
        ws_source = wb_source.active  # Se procesa la hoja activa
    except Exception as e:
        log_message(f"Error al abrir el archivo fuente: {e}")
        messagebox.showerror("Error", f"Error al abrir el archivo fuente: {e}")
        return

    # Abrir archivo destino para modificación
    try:
        log_message("Abriendo archivo destino...")
        wb_dest = openpyxl.load_workbook(destination_file)
        ws_dest = wb_dest.active  # Se usa la hoja activa del destino
    except Exception as e:
        log_message(f"Error al abrir el archivo destino: {e}")
        messagebox.showerror("Error", f"Error al abrir el archivo destino: {e}")
        wb_source.close()
        return

    # Buscar la primera fila vacía en la columna C del destino (suponiendo encabezados en la fila 1)
    dest_row = 2
    while ws_dest.cell(dest_row, 3).value is not None:
        dest_row += 1
    log_message(f"Iniciando copia en la fila {dest_row} del archivo destino.")

    # Recopilar Part Numbers existentes en la columna C del destino para evitar duplicados
    existing_part_numbers = set()
    for (pn,) in ws_dest.iter_rows(min_row=2, min_col=3, max_col=3, values_only=True):
        if pn:
            existing_part_numbers.add(str(pn).strip())

    copied_rows = 0
    # Iterar sobre todas las filas del archivo fuente (omitiendo encabezados)
    for row in ws_source.iter_rows(min_row=2):
        # Solo se procesa la fila si el valor de la columna A (índice 0) coincide con allowed_categories
        category = str(row[0].value or "").strip()
        if category not in allowed_categories:
            log_message(f"Fila ignorada. Categoría '{category}' no permitida.")
            continue

        # Obtener el Part Number desde la columna C (índice 2)
        part_number = str(row[2].value or "").strip()
        if not part_number:
            log_message("Fila sin Part Number, se omite.")
            continue

        # Evitar duplicados en el destino
        if part_number in existing_part_numbers:
            log_message(f"Se omite Part Number duplicado: {part_number}")
            continue

        # Obtener el precio desde la columna L (índice 11) y verificar que no sea 0
        price = row[11].value
        try:
            if float(price) == 0:
                log_message(f"Se omite el item {part_number} por tener precio 0.")
                continue
        except (ValueError, TypeError):
            log_message(f"Se omite el item {part_number} debido a precio no numérico.")
            continue

        # Obtener el valor de la columna S (índice 18)
        value_S = row[18].value

        # Escribir en el destino:
        # - Columna C y D: Part Number
        ws_dest.cell(dest_row, 3, part_number)
        ws_dest.cell(dest_row, 4, part_number)
        # - Columna E: Valor de la columna S del fuente
        ws_dest.cell(dest_row, 5, value_S)
        # - Columna H: Precio (valor de la columna L del fuente)
        ws_dest.cell(dest_row, 8, price)

        log_message(f"Insertado Part Number {part_number} en la fila {dest_row}.")
        existing_part_numbers.add(part_number)
        dest_row += 1
        copied_rows += 1

    try:
        wb_dest.save(destination_file)
        log_message(f"Operación completada. Filas copiadas: {copied_rows}.")
        messagebox.showinfo("Éxito", "Datos copiados exitosamente.")
    except PermissionError as pe:
        log_message(f"Error de permisos: {pe}. Cierra el archivo destino e intenta de nuevo.")
        messagebox.showerror("Error", f"Error de permisos: {pe}")
    except Exception as e:
        log_message(f"Error al guardar el archivo destino: {e}")
        messagebox.showerror("Error", f"Error al guardar el archivo destino: {e}")
    finally:
        wb_source.close()
        wb_dest.close()

def main():
    global log_text, source_file, destination_file
    source_file = None
    destination_file = None

    root = tk.Tk()
    root.title("Copiador Futurista de Excel")
    root.geometry("600x520")

    frame = tk.Frame(root)
    frame.pack(pady=10)

    tk.Button(frame, text="Seleccionar archivo fuente", command=open_file).pack(side=tk.LEFT, padx=5)
    tk.Button(frame, text="Seleccionar archivo destino", command=select_destination).pack(side=tk.LEFT, padx=5)

    tk.Button(root, text="Copiar Datos", command=copy_data).pack(pady=15)

    tk.Label(root, text="Logs de ejecución:").pack()
    global log_text
    log_text = tk.Text(root, height=15)
    log_text.pack(pady=10, fill=tk.BOTH, expand=True)

    root.mainloop()

if __name__ == "__main__":
    main()
